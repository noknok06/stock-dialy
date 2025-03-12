# company_master/admin.py
from django.contrib import admin
from django.http import HttpResponseRedirect
from django.urls import path
from django.shortcuts import render
from django import forms
from .models import CompanyMaster
import pandas as pd
import numpy as np
import re
from django.contrib import messages
from django.db import transaction
from datetime import datetime

class ExcelUploadForm(forms.Form):
    """管理画面用のExcelアップロードフォーム"""
    excel_file = forms.FileField(
        label='Excelファイル',
        help_text='SBI証券などから取得したデータファイルをアップロードしてください'
    )
    replace_existing = forms.BooleanField(
        label='既存データを置き換える',
        required=False,
        initial=True,
        help_text='チェックすると、既存の企業マスタを全て削除して新しいデータで置き換えます。チェックしない場合は、既存データに追加または更新されます。'
    )

@admin.register(CompanyMaster)
class CompanyMasterAdmin(admin.ModelAdmin):
    """企業マスタの管理画面クラス"""
    list_display = ('code', 'name', 'market', 'industry_name_33', 'scale_name', 'unit')
    list_filter = ('market', 'industry_name_33', 'scale_name')
    search_fields = ('code', 'name')
    readonly_fields = ('created_at', 'updated_at')
    ordering = ('code',)
    
    # カスタムアクションの追加
    actions = ['export_to_excel']
    
    # カスタムURLの追加
    def get_urls(self):
        urls = super().get_urls()
        my_urls = [
            path('import-excel/', self.admin_site.admin_view(self.import_excel), name='company_master_import_excel'),
        ]
        return my_urls + urls
    
    # 変更部分: 独自のリンクを管理画面に追加
    def changelist_view(self, request, extra_context=None):
        extra_context = extra_context or {}
        extra_context['show_import_button'] = True
        return super().changelist_view(request, extra_context=extra_context)

    
    def import_excel(self, request):
        """Excelファイルから企業データをインポートするビュー"""
        if request.method == 'POST':
            form = ExcelUploadForm(request.POST, request.FILES)
            if form.is_valid():
                excel_file = form.cleaned_data['excel_file']
                replace_existing = form.cleaned_data['replace_existing']
                
                try:
                    # ExcelをPandasで読み込み
                    df = pd.read_excel(excel_file, header=None)
                    
                    # カラム名を明示的に設定
                    df.columns = [
                        '日付', 'コード', '銘柄名', '市場・商品区分', 
                        '33業種コード', '33業種区分', 
                        '17業種コード', '17業種区分', 
                        '規模コード', '規模区分'
                    ]
                    
                    # 日付を YYYY-MM-DD 形式に変換
                    def convert_date(date_str):
                        try:
                            # 文字列から日付オブジェクトに変換
                            return datetime.strptime(str(date_str), '%Y%m%d').strftime('%Y-%m-%d')
                        except (ValueError, TypeError):
                            # 変換できない場合は None を返す
                            return None
                    
                    # 日付変換
                    df['日付'] = df['日付'].apply(convert_date)
                    
                    # 証券コードを整形（数値の場合文字列に変換し、4桁になるよう0埋め）
                    df['コード'] = df['コード'].apply(lambda x: 
                        str(x).zfill(4) if pd.notna(x) and str(x).strip() and re.match(r'^\d+$', str(x).strip()) 
                        else str(x).strip() if pd.notna(x) else '')
                    
                    # NaNを空文字列に変換
                    for col in df.columns:
                        if col != '日付':
                            df[col] = df[col].fillna('').astype(str)
                    
                    # トランザクションを開始
                    with transaction.atomic():
                        if replace_existing:
                            # 既存データを全て削除
                            CompanyMaster.objects.all().delete()
                        
                        # 新しいデータを作成するためのリスト
                        companies_to_create = []
                        update_count = 0
                        
                        # 既存レコードのコードを取得
                        existing_codes = set(CompanyMaster.objects.values_list('code', flat=True))
                        
                        for _, row in df.iterrows():
                            # 必須フィールドが欠けている場合はスキップ
                            if not row['コード'] or not row['銘柄名']:
                                continue
                            
                            # 企業データを作成
                            company_data = {
                                'date': row['日付'],
                                'code': row['コード'],
                                'name': row['銘柄名'],
                                'market': row['市場・商品区分'],
                                'industry_code_33': row['33業種コード'],
                                'industry_name_33': row['33業種区分'],
                                'industry_code_17': row['17業種コード'],
                                'industry_name_17': row['17業種区分'],
                                'scale_code': row['規模コード'],
                                'scale_name': row['規模区分'],
                                'unit': 100  # デフォルト値
                            }
                            
                            # 既存レコードは更新、新規レコードは作成
                            if company_data['code'] in existing_codes:
                                CompanyMaster.objects.filter(code=company_data['code']).update(**company_data)
                                update_count += 1
                            else:
                                companies_to_create.append(CompanyMaster(**company_data))
                        
                        # バルク作成で効率化
                        if companies_to_create:
                            CompanyMaster.objects.bulk_create(companies_to_create)
                        
                        # 結果メッセージを表示
                        total_created = len(companies_to_create)
                        total_processed = total_created + update_count
                        
                        messages.success(request, 
                            f'インポートが完了しました。{total_processed}件のデータを処理しました。'
                            f'（新規: {total_created}件、更新: {update_count}件）'
                        )
                
                except Exception as e:
                    messages.error(request, f'エラーが発生しました: {str(e)}')
                
                return HttpResponseRedirect("../")
        else:
            form = ExcelUploadForm()
        
        context = {
            'form': form,
            'title': 'Excelからの企業マスタインポート',
            'opts': self.model._meta,
        }
        return render(request, 'admin/company_master/import_excel.html', context)
    
    def export_to_excel(self, request, queryset):
        """選択した企業データをExcelにエクスポート"""
        import io
        from django.http import HttpResponse
        from openpyxl import Workbook
        
        output = io.BytesIO()
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = '企業マスタ'
        
        # ヘッダー行
        headers = [
            '日付', 'コード', '銘柄名', '市場・商品区分', 
            '33業種コード', '33業種区分', 
            '17業種コード', '17業種区分', 
            '規模コード', '規模区分', '売買単位'
        ]
        for col_num, header in enumerate(headers, 1):
            worksheet.cell(row=1, column=col_num, value=header)
        
        # データ行
        for row_num, company in enumerate(queryset, 2):
            worksheet.cell(row=row_num, column=1, value=company.date)
            worksheet.cell(row=row_num, column=2, value=company.code)
            worksheet.cell(row=row_num, column=3, value=company.name)
            worksheet.cell(row=row_num, column=4, value=company.market)
            worksheet.cell(row=row_num, column=5, value=company.industry_code_33)
            worksheet.cell(row=row_num, column=6, value=company.industry_name_33)
            worksheet.cell(row=row_num, column=7, value=company.industry_code_17)
            worksheet.cell(row=row_num, column=8, value=company.industry_name_17)
            worksheet.cell(row=row_num, column=9, value=company.scale_code)
            worksheet.cell(row=row_num, column=10, value=company.scale_name)
            worksheet.cell(row=row_num, column=11, value=company.unit)
        
        workbook.save(output)
        output.seek(0)
        
        # レスポンスの作成
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=company_master.xlsx'
        
        return response
    
    export_to_excel.short_description = '選択した企業をExcelにエクスポート'